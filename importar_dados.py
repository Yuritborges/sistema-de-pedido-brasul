"""
importar_dados.py
Execute UMA VEZ na pasta raiz do projeto:
    python importar_dados.py

Lê as planilhas Excel e gera obras.json e fornecedores.json em assets/
"""
import os, json
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ASSETS   = os.path.join(BASE_DIR, "assets")
os.makedirs(ASSETS, exist_ok=True)

# ── Configure aqui os caminhos das suas planilhas ────────────────────────────
PLANILHAS = [
    # Coloque os caminhos completos ou relativos das planilhas
    "PEDIDOS_IURY_-_ABRIL_2026.xlsx",
    "PEDIDOS_THAMYRES__xlsb.xlsx",
]

obras        = {}
fornecedores = {}

for arq in PLANILHAS:
    if not os.path.exists(arq):
        print(f"[AVISO] Arquivo não encontrado: {arq}")
        continue

    print(f"[LENDO] {arq}...")
    wb = openpyxl.load_workbook(arq, read_only=True, data_only=True)

    # ── Obras ─────────────────────────────────────────────────────────────────
    if "DADOS DAS OBRAS" in wb.sheetnames:
        ws = wb["DADOS DAS OBRAS"]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0: continue  # pula cabeçalho
            nome = str(row[0]).strip() if row[0] else ""
            if not nome or nome == "None": continue
            if nome not in obras:  # mantém o primeiro encontrado
                obras[nome] = {
                    "faturamento": str(row[1]).strip() if row[1] else "BRASUL",
                    "escola":      str(row[2]).strip() if row[2] else "",
                    "endereco":    str(row[3]).strip() if row[3] else "",
                    "bairro":      str(row[4]).strip() if row[4] else "",
                    "cep":         str(row[5]).strip() if row[5] else "",
                    "contrato":    str(row[6]).strip() if row[6] else "0",
                    "cidade":      str(row[7]).strip() if row[7] else "",
                    "uf":          str(row[8]).strip() if row[8] else "SP",
                    "empreiteiro": str(row[9]).strip() if row[9] else "",
                    "contato":     str(row[10]).strip() if row[10] else "",
                }

    # ── Fornecedores ──────────────────────────────────────────────────────────
    if "FORNECEDORES" in wb.sheetnames:
        ws = wb["FORNECEDORES"]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0: continue
            nome = str(row[0]).strip() if row[0] else ""
            if not nome or nome == "None": continue
            if nome not in fornecedores:
                fornecedores[nome] = {
                    "razao":      str(row[1]).strip() if row[1] else "",
                    "email":      str(row[2]).strip() if row[2] else "",
                    "vendedor":   str(row[3]).strip() if row[3] else "",
                    "telefone":   str(row[4]).strip() if row[4] else "",
                    "pix":        str(row[5]).strip() if row[5] else "",
                    "favorecido": str(row[6]).strip() if row[6] else "",
                }

    wb.close()

# ── Salva os JSONs ────────────────────────────────────────────────────────────
saida_obras = os.path.join(ASSETS, "obras.json")
saida_forns = os.path.join(ASSETS, "fornecedores.json")

with open(saida_obras, "w", encoding="utf-8") as f:
    json.dump(obras, f, ensure_ascii=False, indent=2)

with open(saida_forns, "w", encoding="utf-8") as f:
    json.dump(fornecedores, f, ensure_ascii=False, indent=2)

print(f"\n✅ {len(obras)} obras salvas em: {saida_obras}")
print(f"✅ {len(fornecedores)} fornecedores salvos em: {saida_forns}")
print("\nImportação concluída! Abra o programa normalmente.")
